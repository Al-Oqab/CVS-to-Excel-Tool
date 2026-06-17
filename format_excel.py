import pandas as pd
import os
import sys
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.marker import DataPoint
from fpdf import FPDF  

def generate_audit_report(input_csv: str, output_excel: str, output_image: str, output_pdf: str):
    if not os.path.exists(input_csv):
        print(f"Error: The file '{input_csv}' was not found.")
        sys.exit(1)

    print(f"Reading data from {input_csv}...")
    df = pd.read_csv(input_csv)
    wb = Workbook()

    col_names = list(df.columns)
    decision_col_name = next((c for c in ["AI Decision", "Status", "Decision"] if c in col_names), None)
    decision_col_idx = col_names.index(decision_col_name) if decision_col_name is not None else None
    wrap_cols = {"Reason", "Notes", "Comment"}
    link_cols = {"MilesLink"}

    # ==========================================
    # 1. EXCEL: AUDIT DATA SHEET
    # ==========================================
    ws_data = wb.active
    ws_data.title = "Audit Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    thin_side = Side(style='thin', color='D9D9D9')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = ws_data.dimensions

    col_width_map = {
        'Timestamp': 22, 'RunID': 18, 'ReportID': 13, 'VehicleID': 13,
        'LicensePlate': 16, 'AI Decision': 16, 'Status': 16,
        'Reason': 55, 'NewDamage': 13, 'MilesLink': 65,
    }
    for i, name in enumerate(col_names, 1):
        letter = ws_data.cell(row=1, column=i).column_letter
        ws_data.column_dimensions[letter].width = col_width_map.get(name, 15)

    pastel_colors = {
        "NEW_DAMAGE": "E2F0D9",
        "DISCARD": "FFF2CC",
        "ERROR": "FCE4D6",
        "MANUAL_REVIEW": "DDEBF7"
    }
    text_colors = {
        "NEW_DAMAGE": "385723",
        "DISCARD": "7F6000",
        "ERROR": "C00000",
        "MANUAL_REVIEW": "1F4E78"
    }

    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        status = str(row[decision_col_idx].value).strip() if decision_col_idx is not None and row[decision_col_idx].value else ""
        fill_color = pastel_colors.get(status, "FFFFFF")
        text_color = text_colors.get(status, "000000")

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(name="Arial", size=10, color=text_color, bold=(status in pastel_colors))

        for cell in row:
            cell.border = thin_border
            col_name = col_names[cell.column - 1] if cell.column - 1 < len(col_names) else ""
            if col_name in wrap_cols:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_name in link_cols and cell.value:
                url = str(cell.value).strip()
                cell.hyperlink = url
                cell.value = url
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if decision_col_idx is not None:
            row[decision_col_idx].fill = fill
            row[decision_col_idx].font = font

    # ==========================================
    # 2. EXCEL: EXECUTIVE DASHBOARD & CHART
    # ==========================================
    print("Generating Executive Dashboard...")
    ws_summary = wb.create_sheet(title="Dashboard")
    ws_summary.sheet_view.showGridLines = False

    ws_summary["A2"] = "QA Audit Executive Summary"
    ws_summary["A2"].font = Font(name="Arial", size=18, bold=True, color="1F4E78")

    total_records = len(df)
    total_errors = len(df[df[decision_col_name] == "ERROR"]) if decision_col_name else 0
    gdpr_discards = len(df[df["Reason"].str.contains("GDPR|Privacy", case=False, na=False)]) if "Reason" in df.columns else 0

    def create_kpi_card(ws, start_col, label, value, bg_color, text_color):
        c1, c2 = start_col, chr(ord(start_col) + 1)
        ws.merge_cells(f"{c1}4:{c2}4")
        ws.merge_cells(f"{c1}5:{c2}5")
        ws[f"{c1}4"] = label
        ws[f"{c1}5"] = value
        
        card_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        lbl_font = Font(name="Arial", size=10, color=text_color, bold=True)
        val_font = Font(name="Arial", size=20, color=text_color, bold=True)
        
        for r in ["4", "5"]:
            ws[f"{c1}{r}"].fill = card_fill
            ws[f"{c2}{r}"].fill = card_fill
            ws[f"{c1}{r}"].alignment = Alignment(horizontal="center", vertical="center")
            
        ws[f"{c1}4"].font = lbl_font
        ws[f"{c1}5"].font = val_font

    create_kpi_card(ws_summary, "A", "TOTAL AUDITED", total_records, "1F4E78", "FFFFFF")
    create_kpi_card(ws_summary, "D", "AI ERRORS", total_errors, "FCE4D6", "C00000")
    create_kpi_card(ws_summary, "G", "GDPR PRIVACY DISCARDS", gdpr_discards, "FFF2CC", "7F6000")

    if decision_col_name:
        counts = df[decision_col_name].value_counts()

        ws_summary["A8"] = "Decision Breakdown"
        ws_summary["A8"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")
        ws_summary.append([])
        ws_summary.append([decision_col_name, "Count"]) 
        
        category_order = []
        for decision, count in counts.items():
            ws_summary.append([decision, count])
            category_order.append(decision)

        for cell in ws_summary[10]:
            if cell.column_letter in ["A", "B"]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
        start_row = 11
        end_row = start_row + len(counts) - 1
        for row in ws_summary.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.column_dimensions['A'].width = 22
        ws_summary.column_dimensions['B'].width = 12

        pie = PieChart()
        labels = Reference(ws_summary, min_col=1, min_row=start_row, max_row=end_row)
        data_ref = Reference(ws_summary, min_col=2, min_row=start_row - 1, max_row=end_row)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribution of AI Decisions"
        pie.width = 16
        pie.height = 11
        
        series = pie.series[0]
        for i, cat in enumerate(category_order):
            if cat in pastel_colors:
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = pastel_colors[cat]
                series.dPt.append(pt)

        ws_summary.add_chart(pie, "D8")

        # ==========================================
        # 3. STANDALONE IMAGE GENERATION
        # ==========================================
        print("Exporting chart as a standalone image...")
        mpl_colors_map = {k: f"#{v}" for k, v in pastel_colors.items()}
        plot_colors = [mpl_colors_map.get(cat, "#CCCCCC") for cat in category_order]

        plt.figure(figsize=(8, 6))
        plt.pie(counts, labels=category_order, autopct='%1.1f%%', startangle=140, 
                colors=plot_colors, textprops={'fontsize': 12}, 
                wedgeprops={'edgecolor': 'gray', 'linewidth': 0.5})
        plt.title("Distribution of AI Decisions", fontsize=16, fontweight='bold', color='#1F4E78')
        plt.axis('equal') 
        
        plt.savefig(output_image, bbox_inches='tight', transparent=False, dpi=300)
        plt.close()

    wb.save(output_excel)
    print(f"Success! Excel report generated at: {output_excel}")



if __name__ == "__main__":
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    INPUT_DIR = os.path.join(BASE_DIR, "input")
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")

    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    csv_files = [f for f in os.listdir(INPUT_DIR) if f.endswith(".csv")]

    if not csv_files:
        print(f"Error: No CSV file found in the 'input' folder: {INPUT_DIR}")
        sys.exit(1)
    if len(csv_files) > 1:
        print(f"Error: Multiple CSV files found: {csv_files}. Please keep only one in the 'input' folder.")
        sys.exit(1)

    csv_name = os.path.splitext(csv_files[0])[0]
    INPUT_FILE = os.path.join(INPUT_DIR, csv_files[0])
    OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, f"{csv_name}.xlsx")
    OUTPUT_IMAGE = os.path.join(OUTPUT_DIR, "Decisions_Chart_v3.png")
    OUTPUT_PDF = os.path.join(OUTPUT_DIR, f"{csv_name}.pdf")

    generate_audit_report(INPUT_FILE, OUTPUT_EXCEL, OUTPUT_IMAGE, OUTPUT_PDF)